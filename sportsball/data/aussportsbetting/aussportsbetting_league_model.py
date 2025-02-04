"""Aussportsbetting league model."""

# pylint: disable=line-too-long
from io import BytesIO
from typing import Any, Iterator

import requests_cache
import tqdm
from dateutil.parser import parse
from openpyxl import load_workbook

from ..game_model import GameModel
from ..league import League
from ..league_model import LeagueModel
from .aussportsbetting_game_model import create_aussportsbetting_game_model


class AusSportsBettingLeagueModel(LeagueModel):
    """AusSportsBetting implementation of the league model."""

    def __init__(
        self,
        league: League,
        session: requests_cache.CachedSession,
        position: int | None = None,
    ) -> None:
        super().__init__(league, session, position=position)
        match league:
            case League.AFL:
                self._spreadsheet_url = (
                    "https://www.aussportsbetting.com/historical_data/afl.xlsx"
                )
            case League.NFL:
                self._spreadsheet_url = (
                    "https://www.aussportsbetting.com/historical_data/nfl.xlsx"
                )
            case _:
                raise ValueError(
                    f"League {league} not supported by aus sports betting."
                )

    def _row_to_game(self, row: Any) -> GameModel | None:
        current_cell_idx = 0
        date_cell = str(row[0].value)
        if date_cell in {"Date", "None"}:
            return None
        current_cell_idx += 1
        if self.league == League.AFL:
            time_cell = str(row[current_cell_idx].value)
            dt = parse(" ".join([date_cell, time_cell]))
            current_cell_idx += 1
        else:
            dt = parse(date_cell)
        home_team = str(row[current_cell_idx].value).strip()
        current_cell_idx += 1
        away_team = str(row[current_cell_idx].value).strip()
        current_cell_idx += 1
        venue = None
        if self.league == League.AFL:
            venue = str(row[current_cell_idx].value).strip()
            current_cell_idx += 1
        home_points = float(row[current_cell_idx].value)  # type: ignore
        current_cell_idx += 1
        away_points = float(row[current_cell_idx].value)  # type: ignore
        current_cell_idx += 1

        if self.league == League.NFL:
            current_cell_idx += 1
        play_off = row[current_cell_idx].value == "Y"

        if self.league == League.NFL:
            current_cell_idx += 2
        else:
            current_cell_idx += 4
        home_odds = float(row[current_cell_idx].value)  # type: ignore
        current_cell_idx += 1
        if self.league == League.NFL:
            current_cell_idx += 3
        away_odds = float(row[current_cell_idx].value)  # type: ignore
        current_cell_idx += 1
        return create_aussportsbetting_game_model(
            dt,
            home_team,
            away_team,
            venue,
            self.session,
            home_points,
            away_points,
            home_odds,
            away_odds,
            self.league,
            play_off,
        )

    @property
    def games(self) -> Iterator[GameModel]:
        with self.session.cache_disabled():
            response = self.session.get(self._spreadsheet_url)
        response.raise_for_status()
        workbook = load_workbook(filename=BytesIO(response.content))
        ws = workbook.active
        if ws is None:
            raise ValueError("ws is null.")
        with tqdm.tqdm(position=self.position) as pbar:
            for row in ws.iter_rows():
                game_model = self._row_to_game(row)
                if game_model is not None:
                    pbar.update(1)
                    pbar.set_description(f"AusSportsBetting - {game_model.dt}")
                    yield game_model
