"""AFL aussportsbetting league model."""

from io import BytesIO
from typing import Iterator

import requests
import tqdm
from dateutil.parser import parse
from openpyxl import load_workbook

from ...game_model import GameModel
from ...league import League
from ...league_model import LeagueModel
from .afl_aussportsbetting_game_model import \
    create_afl_aussportsbetting_game_model


class AFLAusSportsBettingLeagueModel(LeagueModel):
    """AFL AusSportsBetting implementation of the league model."""

    def __init__(self, session: requests.Session) -> None:
        super().__init__(League.AFL, session)

    @property
    def games(self) -> Iterator[GameModel]:
        response = self.session.get(
            "https://www.aussportsbetting.com/historical_data/afl.xlsx"
        )
        response.raise_for_status()
        workbook = load_workbook(filename=BytesIO(response.content))
        ws = workbook.active
        if ws is None:
            raise ValueError("ws is null.")
        for row in tqdm.tqdm(ws.iter_rows(), desc="AusSportsBetting Games"):
            date_cell = str(row[0].value)
            if date_cell in {"Date", "None"}:
                continue
            time_cell = str(row[1].value)
            dt = parse(" ".join([date_cell, time_cell]))
            home_team = str(row[2].value).strip()
            away_team = str(row[3].value).strip()
            venue = str(row[4].value).strip()
            home_points = float(row[5].value)  # type: ignore
            away_points = float(row[6].value)  # type: ignore
            home_odds = float(row[12].value)  # type: ignore
            away_odds = float(row[13].value)  # type: ignore
            yield create_afl_aussportsbetting_game_model(
                dt,
                home_team,
                away_team,
                venue,
                self.session,
                home_points,
                away_points,
                home_odds,
                away_odds,
                self.league,
            )
