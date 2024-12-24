"""AFL aussportsbetting odds model."""

import datetime
from io import BytesIO

import requests
from dateutil.parser import parse
from openpyxl import Workbook, load_workbook

from ...team_model import OddsModel
from .afl_aussportsbetting_bookie_model import \
    create_afl_aussportsbetting_bookie_model

_ODDS_XLSX_URL = "https://www.aussportsbetting.com/historical_data/afl.xlsx"
_WORKBOOK: Workbook | None = None
_TEAM_NAME_TRANSLATIONS = {
    "Greater Western Sydney": "GWS Giants",
    "Brisbane Lions": "Brisbane",
}


def create_afl_aussportsbetting_odds_model(
    session: requests.Session, date: datetime.date, team_name: str
) -> OddsModel:
    """Create an odds model based off aus sports betting."""
    # pylint: disable=global-statement
    global _WORKBOOK
    team_name = _TEAM_NAME_TRANSLATIONS.get(team_name, team_name)
    if _WORKBOOK is None:
        response = session.get(_ODDS_XLSX_URL)
        _WORKBOOK = load_workbook(filename=BytesIO(response.content))
    ws = _WORKBOOK.active
    if ws is None:
        raise ValueError("ws is null.")
    odds = None
    for row in ws.iter_rows():
        date_cell = str(row[0].value)
        if date_cell in {"Date", "None"}:
            continue
        potential_date = parse(date_cell).date()
        if date != potential_date:
            continue
        home_team = str(row[2].value).strip()
        if home_team == team_name:
            odds = float(str(row[12].value))
            break
        away_team = str(row[3].value).strip()
        if away_team == team_name:
            odds = float(str(row[13].value))
            break
    if odds is None:
        raise ValueError(f"odds is null with date {date} team_name {team_name}.")
    return OddsModel(odds=odds, bookie=create_afl_aussportsbetting_bookie_model())
